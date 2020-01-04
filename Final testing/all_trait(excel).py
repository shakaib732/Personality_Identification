
from openpyxl import Workbook
from openpyxl import load_workbook
wb1 =load_workbook("E:/New folder (2)/new.xlsx")
wb2 = load_workbook("E:/New folder (2)/essay.xls.xlsx")
ws1 = wb1.active
ws2 = wb2.active
i=2
sheet1=wb2.get_sheet_by_name('essays')
for x in range(2,2469):#(for extraversion)
    if ((sheet1.cell(row=x,column=3).value=='y') and (sheet1.cell(row=x,column=4).value=='n') and (sheet1.cell(row=x,column=5).value=='n') and (sheet1.cell(row=x,column=6).value=='n') and (sheet1.cell(row=x,column=7).value=='n')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=3).value=(text)
            i=i+1
i=2
for x in range(2,2469):#(for neuroticism)
    if ((sheet1.cell(row=x,column=3).value=='n') and (sheet1.cell(row=x,column=4).value=='y') and (sheet1.cell(row=x,column=5).value=='n') and (sheet1.cell(row=x,column=6).value=='n') and (sheet1.cell(row=x,column=7).value=='n')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=5).value=(text)
            i=i+1

i=2
for x in range(2,2469):#(for agreableness)
    if ((sheet1.cell(row=x,column=3).value=='n') and (sheet1.cell(row=x,column=4).value=='n') and (sheet1.cell(row=x,column=5).value=='y') and (sheet1.cell(row=x,column=6).value=='n') and (sheet1.cell(row=x,column=7).value=='n')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=4).value=(text)
            i=i+1

i=2
for x in range(2,2469):#(for consitiousness)
    if ((sheet1.cell(row=x,column=3).value=='n') and (sheet1.cell(row=x,column=4).value=='n') and (sheet1.cell(row=x,column=5).value=='n') and (sheet1.cell(row=x,column=6).value=='y') and (sheet1.cell(row=x,column=7).value=='n')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=2).value=(text)
            i=i+1

i=2
for x in range(2,2469):#(for openess)
    if ((sheet1.cell(row=x,column=3).value=='n') and (sheet1.cell(row=x,column=4).value=='n') and (sheet1.cell(row=x,column=5).value=='n') and (sheet1.cell(row=x,column=6).value=='n') and (sheet1.cell(row=x,column=7).value=='y')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=1).value=(text)
            i=i+1

i=2
for x in range(2,2469):#(for none)
    if ((sheet1.cell(row=x,column=3).value=='n') and (sheet1.cell(row=x,column=4).value=='n') and (sheet1.cell(row=x,column=5).value=='n') and (sheet1.cell(row=x,column=6).value=='n') and (sheet1.cell(row=x,column=7).value=='n')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=6).value=(text)
            i=i+1

i=2
for x in range(2,2469):#(for all)
    if ((sheet1.cell(row=x,column=3).value=='y') and (sheet1.cell(row=x,column=4).value=='y') and (sheet1.cell(row=x,column=5).value=='y') and (sheet1.cell(row=x,column=6).value=='y') and (sheet1.cell(row=x,column=7).value=='y')):
            text=sheet1.cell(row=x, column=2).value
            ws1.cell(row=i,column=7).value=(text)
            i=i+1
            
wb1.save("E:/New folder (2)/new.xlsx")
print('done')
