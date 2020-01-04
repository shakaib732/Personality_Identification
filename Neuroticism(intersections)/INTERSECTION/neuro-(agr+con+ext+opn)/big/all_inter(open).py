import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook("E:/New folder (2)/Neuroticism(intersections)/INTERSECTION/neuro-(agr+con+ext+opn)/neuro.xlsx")
ws=wb.active
with open("E:\\New folder (2)\\Neuroticism(intersections)\\INTERSECTION\\fn_inte_bigram\\neuro-const.txt", 'rb') as p:
        a = pickle.load(p)
        
with open("E:\\New folder (2)\\Neuroticism(intersections)\\INTERSECTION\\fn_inte_bigram\\neuro-agreb.txt", 'rb') as q:
    b = pickle.load(q)
    
with open("E:\\New folder (2)\\Neuroticism(intersections)\\INTERSECTION\\fn_inte_bigram\\neuro-extra.txt", 'rb') as r:
    c = pickle.load(r)
    
with open("E:\\New folder (2)\\Neuroticism(intersections)\\INTERSECTION\\fn_inte_bigram\\neuro-open.txt", 'rb') as s:
    d = pickle.load(s)
    
#g=list(set(a) & set(b) & set(c) & set(d))
g=[val for val in a if val in b and val in c and val in d]
print(g)
with open("E:\\New folder (2)\\testing\\neuro\\bi_inter.txt", 'wb') as f:
    pickle.dump(g, f)
h=len(g)
print(h)

ws.cell(row=1,column=2).value=("BIGRAM")
i=0
for x in range(2,h):
    ws.cell(row=x,column=2).value=("%s %s"%g[i])
    i+=1

ws.cell(row=h,column=2).value=("%s %s"%g[h-1])
wb.save("E:/New folder (2)/Neuroticism(intersections)/INTERSECTION/neuro-(agr+con+ext+opn)/neuro.xlsx")

