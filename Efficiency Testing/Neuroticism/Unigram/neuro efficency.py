import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
wb1 =load_workbook("E:/New folder (2)/new.xlsx")
ws1 = wb1.active
sheet=wb1.get_sheet_by_name('Sheet')
list1=list2=[]

with open("E:\\New folder (2)\\testing\\open\\uni_inter.txt", 'rb') as m:
    opn = pickle.load(m)
    p=len(opn)

with open("E:\\New folder (2)\\testing\\const\\uni_inter.txt", 'rb') as m:
    con = pickle.load(m)
    q=len(con)

with open("E:\\New folder (2)\\testing\\extra\\uni_inter.txt", 'rb') as m:
    ext = pickle.load(m)
    r=len(ext)

with open("E:\\New folder (2)\\testing\\agreb\\uni_inter.txt", 'rb') as m:
    agr = pickle.load(m)
    s1=len(agr)

with open("E:\\New folder (2)\\testing\\neuro\\uni_inter.txt", 'rb') as m:
    neu = pickle.load(m)
    t=len(neu)
Copn=Ccon=Cagr=Cext=Cneu=0
openn=const=agreb=extra=neuro=0
o_count=c_count=e_count=a_count=n_count=maximum=maximum2=0
o_final=c_final=e_final=a_final=n_final=0
o_=c_=e_=a_=n_=0

def splitParagraphIntoSentences(paragraph):
    
    import re
    
    sentenceEnders = re.compile('[.]')
    sentenceList = sentenceEnders.split(paragraph)
    return sentenceList

'''def hige(a,b,c,d,e):
    
    rid=[]
    rid.append(a)
    rid.append(b)
    rid.append(c)
    rid.append(d)
    rid.append(e)
    rid1=set()
    rid1.add(a)
    rid1.add(b)
    rid1.add(c)
    rid1.add(d)
    rid1.add(e)
    if(len(rid)==len(rid1)):
        return 0
    else:
        return 1'''
    
list1=[]

for a in range(2,142): #for open
    text=sheet.cell(row=a,column=5).value
    s=splitParagraphIntoSentences(text)
    list1=s
    h=len(list1)
    #print(h)
    for b in range(1,h):
        st=list1[b]
        sentences=st.split()
        list3=sentences
        #print(list2[3])
        c=len((list3))
        #print(list3)
        for ze in range(1,c):
            str2=list3[ze]
            str1=str2.lower()
            str3=str1.replace(',',' ')
            #print(str3)
            for d in range(0,p):
                if(str3=="%s"%opn[d]):
                    openn+=1
            for d in range(0,q):
                if(str3=="%s"%con[d]):
                    const+=1
            for d in range(0,r):
                if(str3=="%s"%ext[d]):
                    extra+=1
            for d in range(0,s1):
                if(str3=="%s"%agr[d]):
                    agreb+=1
            for d in range(0,t):
                if(str3=="%s"%neu[d]):
                    neuro+=1
        maximum=max(openn,const,extra,agreb,neuro)
        #if(hige(openn,const,extra,agreb,neuro)==0):
        if(maximum==openn):
            o_count+=1
        if(maximum==const):
            c_count+=1
        if(maximum==extra):
            e_count+=1
        if(maximum==agreb):
            a_count+=1
        if(maximum==neuro):
            n_count+=1
        openn=const=agreb=extra=neuro=0
        maximum=0
    maximum2=max(o_count,c_count,e_count,a_count,n_count)
    if(maximum2==o_count):
        o_final+=1
    if(maximum2==c_count):
        c_final+=1
    if(maximum2==e_count):
        e_final+=1
    if(maximum2==a_count):
        a_final+=1
    if(maximum2==n_count):
        n_final+=1
    maximum2=0
    o_count=c_count=e_count=a_count=n_count=0
    
    
    

               
print("open:", o_final)
print("const:", c_final)
print("extra:", e_final)
print("agreb:", a_final)
print("neuro:", n_final)
    
    

wb1.save("E:/New folder (2)/try.xlsx")
             
       
        


