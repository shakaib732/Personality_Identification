import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook("E:/New folder (2)/Extraversion(intersections)/INTERSECTION/agreb-(opn+con+ext+neu)/extra.xlsx")
ws=wb.active

with open("E:\\New folder (2)\\Extraversion(intersections)\\INTERSECTION\\fn_inte_trigram\\extra-const.txt", 'rb') as p:
        a = pickle.load(p)
        
with open("E:\\New folder (2)\\Extraversion(intersections)\\INTERSECTION\\fn_inte_trigram\\extra-agreb.txt", 'rb') as q:
    b = pickle.load(q)
    
with open("E:\\New folder (2)\\Extraversion(intersections)\\INTERSECTION\\fn_inte_trigram\\extra-open.txt", 'rb') as r:
    c = pickle.load(r)
    
with open("E:\\New folder (2)\\Extraversion(intersections)\\INTERSECTION\\fn_inte_trigram\\extra-neuro.txt", 'rb') as s:
    d = pickle.load(s)
    
#g=list(set(a) & set(b) & set(c) & set(d))

g=[val for val in a if val in b and val in c and val in d]
print(g)
with open("E:\\New folder (2)\\testing\\extra\\tri_inter.txt", 'wb') as f:
    pickle.dump(g, f)
h=len(g)
print(h)

ws.cell(row=1,column=3).value=("TRIGRAM")
i=0
for x in range(2,h+1):
    ws.cell(row=x,column=3).value=("%s %s %s"%g[i])
    i+=1
    
ws.cell(row=h+1,column=3).value=("%s %s %s"%g[h-1])

wb.save("E:/New folder (2)/Extraversion(intersections)/INTERSECTION/agreb-(opn+con+ext+neu)/extra.xlsx")



